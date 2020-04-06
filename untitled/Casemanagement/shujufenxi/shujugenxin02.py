import asyncio
from pyppeteer import launch
import random
import requests
import json
import time
import os
import base64
from openpyxl import load_workbook
import openpyxl
import yaml
from logger import Log

log = Log()

class TrickUrlSession(requests.Session):
    '''重写url禁止url进行urlcode'''
    def setUrl(self, url):
        self._trickUrl = url
    def send(self, request, **kwargs):
        if self._trickUrl:
            request.url = self._trickUrl
        return requests.Session.send(self, request, **kwargs)


def screen_size():
    """使用tkinter获取屏幕大小"""
    import tkinter
    tk = tkinter.Tk()
    width = tk.winfo_screenwidth()
    height = tk.winfo_screenheight()
    tk.quit()
    return width, height


def input_time_random():
    return random.randint(100, 151)


async def try_validation(page, distance=308):
    # 将距离拆分成两段，模拟正常人的行为
    distance1 = distance - 10
    distance2 = 10
    btn_position = await page.evaluate('''
        () =>{
        return {
        x: document.querySelector('#nc_1_n1z').getBoundingClientRect().x,
        y: document.querySelector('#nc_1_n1z').getBoundingClientRect().y,
        width: document.querySelector('#nc_1_n1z').getBoundingClientRect().width,
        height: document.querySelector('#nc_1_n1z').getBoundingClientRect().height
        }}
        ''')
    x = btn_position['x'] + btn_position['width'] / 2
    y = btn_position['y'] + btn_position['height'] / 2
    await page.mouse.move(x, y)
    await page.mouse.down()
    await page.mouse.move(x + distance1, y, {'steps': 30})
    await page.waitFor(800)
    await page.mouse.move(x + distance1 + distance2, y, {'steps': 20})
    await page.waitFor(800)
    await page.mouse.up()


#获取token主函数
async def token_main(url,username,password):
    browser = await launch(headless=False, args=['--disable-infobars'])
    page = await browser.newPage()
    await page.goto(url)
    width, height = screen_size()
    # 最大化窗口
    await page.setViewport(viewport={'width': width, 'height':height})
    # 设置浏览器
    await page.setUserAgent('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)'
                                ' Chrome/74.0.3729.169 Safari/537.36')
    # 防止被识别，将webdriver设置为false
    await page.evaluate(
        '''() =>{ Object.defineProperties(navigator,{ webdriver:{ get: () => false } }) }''')
    #登录的用户名和密码
    await page.type('#BX_Layer2 > div > form > div > div > div > input',username, {'delay': input_time_random() - 50})
    await page.type('.BX_Sprite2 .el-form .el-form-item:nth-child(2) .el-form-item__content .el-input__inner',password, {'delay': input_time_random()})
    # 登录时的验证
    normal_login = await page.xpath('//*[@id="BX_Layer2"]/div/form/div[4]/button')
    await normal_login[0].click()
    time.sleep(1)
    # 滑动验证
    await page.waitFor(1000)
    await try_validation(page)
    await page.waitFor(2000)
    #await shibieguanli(page)
    await page.waitFor(1000)
    #await qudaofengkong(page)
    await page.waitFor(2000)
    #res = await page.cookies()
    login_token = await page.evaluate("window.localStorage.getItem('Sense-Token')",force_expr=True)
    login_companyid = await page.evaluate("window.localStorage.getItem('Sense-CompanyId')", force_expr=True)
    #print("获取登录到的token是:",login_token,"获取登录后用户company:",login_companyid)
    #图片保存
    #await page.screenshot({'path': 'shibieguanli.png'})
    log.info("登录后token:{}-- ".format(login_token))
    log.info("登录后companyid:{}-- ".format(login_companyid))
    await page.close()
    login_kwargs = {"token":login_token,"companyid":login_companyid}
    #print(login_kwargs)
    return login_kwargs


#案场列表获取
async def store_list_find(companyid,headers):
    url = 'https://icloud.sensetime.com/senserealty/store/api/sensego/console/v1.0/store/all'
    querystring = {"company_id":companyid}
    res = requests.session()
    #print("进入的headers:",headers,companyid)
    log.info("案场列表获取querystring:{}-- ".format(str(querystring)))
    log.info("案场列表获取headers:{}-- ".format(str(headers)))
    store_name = res.get(url=url, headers=headers, params=querystring)
    res = store_name.json()
    log.info("案场列表获取请求结果result:{}-- ".format(str(res)))
    return res

def get_store_name(**kwargs):
    #返回案场名称（用作文件命名）
    store_id_name = {}
    #print(kwargs,type(kwargs))
    log.info("返回案场输入字典:{}-- ".format(str(kwargs)))
    if isinstance(kwargs,dict):
        store_list_info = kwargs.get('list')
        if store_list_info:
            for i in range(len(store_list_info)):
                store_info_dict = store_list_info[i]
                store_id_name[store_info_dict['store_id']] = store_info_dict['store_name']
        else:
            store_id_name[str(kwargs)] = kwargs
    else:
        store_id_name[str(kwargs)] = kwargs
    log.info("返回案场输出字典:{}-- ".format(str(store_id_name)))
    return store_id_name

#渠道页面
async def get_qudao_page(store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/trade'
    querystring = {"type": "0", "store_id":store_id if store_id  else None}
    headers = headers if headers else None
    log.info("渠道页面输入参数:{}-- ".format(str(querystring)))
    log.info("渠道页面输入请求头:{}-- ".format(str(headers)))
    res = requests.session()
    qudao_req = res.get(url=url, headers=headers, params=querystring)
    qudao_res = qudao_req.json()
    log.info("渠道页面返回结果:{}-- ".format(str(qudao_res)))
    qudao_list = qudao_res.get('list') if qudao_res.get('list') else store_id
    log.info("渠道页面输出结果:{}-- ".format(str(qudao_list)))
    return store_id,qudao_list

#渠道详情数据
def get_qudao_detail_page(id_number,store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/customer-detail/risk/' + store_id + '/api/sensego/console/v1.0/trade/trace'
    querystring = {"id_number":id_number, "store_id":store_id}
    log.info("渠道详情输入参数:--{}-- ".format(str(querystring)))
    log.info("渠道详情输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    qudao_detail_req = res.get(url=url, headers=headers, params=querystring)
    qudao_detail_res = qudao_detail_req.json()
    id_capture_image = qudao_detail_res.get('id_capture_image') if qudao_detail_res.get('id_capture_image') else str(qudao_detail_res)
    qudao_detail_list = qudao_detail_res.get('list') if qudao_detail_res.get('list') else str(qudao_detail_res)
    log.info("渠道详情返回结果:{}-- ".format(str(qudao_detail_res)))
    log.info("渠道详情输出结果1:{}-- ".format(str(id_capture_image)))
    log.info("渠道详情输出结果2:{}-- ".format(str(qudao_detail_list)))
    return id_capture_image,qudao_detail_list

#自主报备页面数据
async def get_baobei_page(store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/report/all'
    querystring = {"store_id":store_id}      #page_index和count是非必填参数
    log.info("自主报备页面输入参数:--{}-- ".format(str(querystring)))
    log.info("自主报备页面输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    baobei_req = res.get(url=url, headers=headers, params=querystring)
    baobei_res = baobei_req.json()
    baobei_list = baobei_res.get('list') if baobei_res.get('list') else store_id
    log.info("自主报备页面返回结果1:{}-- ".format(str(baobei_res)))
    log.info("自主报备页面输出结果2:{}-- ".format(str(baobei_list)))
    return store_id,baobei_list


#报备详细记录查询数据
def get_baobei_detail_page(store_id,_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/customer-detail/report/'+ store_id +'/api/sensego/console/v1.0/report'
    querystring = {"store_id":store_id,"_id":_id}
    log.info("报备详细记录输入参数:--{}-- ".format(str(querystring)))
    log.info("报备详细记录输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    baobei_detail_req = res.get(url=url, headers=headers, params=querystring)
    baobei_detail_res = baobei_detail_req.json()
    #print("自主报备详情",baobei_detail_res)
    baobei_detail_list = baobei_detail_res.get('results') if baobei_detail_res.get('results') else store_id
    id_detail_image = baobei_detail_res.get('image') if baobei_detail_res.get('image') else str(baobei_detail_res)
    log.info("报备详细记录返回结果:{}-- ".format(str(baobei_detail_res)))
    log.info("报备详细记录输出结果1:{}-- ".format(str(id_detail_image)))
    log.info("报备详细记录输出结果2:{}-- ".format(str(baobei_detail_list)))
    #渠道风控详情返回数据
    return id_detail_image,baobei_detail_list

#以图搜图接口
async def get_search_image_page(group_id,face_image,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/realestate/trace'
    data = {"group_id":group_id,"face_image":face_image}
    res = requests.session()
    log.info("以图搜图接口输入参数:--{}-- ".format(str(data)))
    log.info("以图搜图接口输入请求头:--{}-- ".format(str(headers)))
    #print("请求参数",data)
    get_search_image_req = res.post(url=url, headers=headers,data=json.dumps(data))
    get_search_image_res = get_search_image_req.json()
    #print(get_search_image_res)
    get_search_image_list = get_search_image_res.get('results') if get_search_image_res.get('results') else group_id
    #print("以图搜图",get_search_image_list)
    log.info("以图搜图接口返回结果:{}-- ".format(str(get_search_image_res)))
    log.info("以图搜图接口输出结果:{}-- ".format(str(get_search_image_list)))
    return get_search_image_list

#到访统计-总数
async def get_daofang_page_summary(store_id,end_time,start_time,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/peoplecount/summary'
    data = {"store_id":store_id,"end_time":end_time,"start_time":start_time,"ask_id":"2"}
    log.info("输入参数:--{}-- ".format(str(data)))
    log.info("输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    get_daofang_req = res.post(url=url, headers=headers,data=json.dumps(data))
    get_daofang_res = get_daofang_req.json()
    get_daofagn_count = get_daofang_res.get('capture_count')
    get_first_time_count = get_daofang_res.get('first_time_count')
    #print("到访统计人数",get_daofagn_count,"首次到访统计数据",get_first_time_count)
    log.info("返回结果:{}-- ".format(str(get_daofang_res)))
    return get_daofang_res


#到访统计-访客总数
def get_daofang_page_total(store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/peoplecount/total'
    data = {"store_id":store_id}
    log.info("访客总数输入参数:--{}-- ".format(str(data)))
    log.info("访客总数输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    get_daofang_total_req = res.post(url=url, headers=headers,data=json.dumps(data))
    get_daofang_total_res = get_daofang_total_req.json()
    daofang_people_total = get_daofang_total_res.get('people_count') if get_daofang_total_res.get('people_count') else 0
    #print("到访统计人数数据----",daofang_people_total)
    log.info("访客总数返回结果:{}-- ".format(str(get_daofang_total_res)))
    log.info("访客总数输出结果:{}-- ".format(str(daofang_people_total)))
    return daofang_people_total

#到访统计-访客数据详情
def get_daofang_page_trend(store_id,start_time,end_time,headers):
    trend_url = "https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/peoplecount/trend"
    data = {"store_id":store_id,"start_time":str(start_time),"end_time":str(end_time),"ask_id":"0"}
    log.info("访客数据详情输入参数:--{}-- ".format(str(data)))
    log.info("访客数据详情输入请求头:--{}-- ".format(str(headers)))
    res = requests.session()
    get_daofang_trend_req = res.post(url=trend_url,headers=headers,data=json.dumps(data))
    get_daofang_trend_res = get_daofang_trend_req.json()
    daofang_people_trend = get_daofang_trend_res.get('list') if get_daofang_trend_res.get('list') else 0
    #print("到访统计人数数据",daofang_people_trend)
    log.info("访客数据详情返回结果:{}-- ".format(str(get_daofang_trend_res)))
    log.info("访客数据详情输出结果:{}-- ".format(str(daofang_people_trend)))
    return daofang_people_trend

#筛选识别结果查询时间
def get_shibie_time(store_id,start_time,headers):
    this_headers = headers
    log.info("筛选识别结果输入参数:--{}-- ".format(str(store_id)))
    log.info("筛选识别结果输入参数:--{}-- ".format(str(start_time)))
    log.info("筛选识别结果输入请求头:--{}-- ".format(str(headers)))
    total = get_daofang_page_total(store_id,headers)
    if int(total) > 0:
        trend_start_time = start_time
        trend_end_time = start_time + (60*60*24*10)
        get_trend_list = get_daofang_page_trend(store_id,trend_start_time,trend_end_time,this_headers)
        #print("查询到到访数据：",get_trend_list)
        need_data = {}
        get_trend_cursor = 0
        get_trend_length = len(get_trend_list)
        while get_trend_cursor < get_trend_length:
            trend_date = get_trend_list[get_trend_cursor].get("date")
            capture_count = get_trend_list[get_trend_cursor].get("capture_count")
            need_data[trend_date] = capture_count
            get_trend_cursor += 1
        trend_tmp_end_time = max(need_data.keys(),key=(lambda k:need_data[k]))
        res_utime = timeutil(trend_tmp_end_time+" 20:00:00")
        log.info("筛选识别时间输出时间:--{}-- ".format(str(trend_start_time)))
        log.info("筛选识别时间输出时间:--{}-- ".format(str(res_utime)))
        return trend_start_time,res_utime
    else:
        trend_start_time = start_time
        trend_end_time = start_time + (60 * 60 * 24 * 10)
        log.info("筛选识别时间输出时间:--{}-- ".format(str(trend_start_time)))
        log.info("筛选识别时间输出时间:--{}-- ".format(str(trend_end_time)))
        return trend_start_time,trend_end_time

#识别结果--会先发起storedevice请求
async def get_shibie_storedevice(store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/group/storedevice'
    querystring = {"store_id":store_id}
    log.info("storedevice请求输入参数:{}-- ".format(str(querystring)))
    log.info("storedevice请求输入请求头:{}-- ".format(str(headers)))
    res = requests.session()
    store_devicemap = res.get(url=url, headers=headers, params=querystring)
    devicemap = store_devicemap.json()
    store_map_list = devicemap.get('group_device_map') if devicemap.get('group_device_map') else store_id
    #print(store_map_list,'请求的结果类型是',type(store_map_list))
    log.info("storedevice返回结果:--{}-- ".format(str(devicemap)))
    log.info("storedevice输出结果:--{}-- ".format(str(store_map_list)))
    return store_id,store_map_list


#识别结果-获取设备及创建时间
async def get_shibie_store_device(store_id,headers):
    url = 'https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/group/store/device'
    querystring = {"store_id":store_id}
    log.info("获取设备及创建时间请求输入参数:{}-- ".format(str(querystring)))
    log.info("获取设备及创建时间请求输入请求头:{}-- ".format(str(headers)))
    res = requests.session()
    store_device = res.get(url=url, headers=headers, params=querystring)
    device_res = store_device.json()
    store_device_res = device_res.get('list') if device_res.get('list') else store_id
    log.info("获取设备及创建时间返回结果:--{}-- ".format(str(device_res)))
    log.info("获取设备及创建时间输出结果:--{}-- ".format(str(store_device_res)))
    return store_id,store_device_res


#获取识别结果
def get_shibie_detail(group_id,start_time,end_time,device_id,headers):
    limit = 1000
    info_start_time = start_time
    info_end_time = end_time
    info_group_id = group_id
    info_device_id = device_id
    log.info("获取识别结果输入请求头:{}-- ".format(str(headers)))
    geturl = "https://icloud.sensetime.com/senserealty/recognition/api/sensego/console/v1.0/datacenter/visitinfo/query/device?start_time=%s&end_time=%s&group_id=%s&device_id=%s&limit=%s" % (
    info_start_time, info_end_time, info_group_id, info_device_id, limit)
    #print(geturl)
    log.info("获取识别结果请求url:{}-- ".format(str(geturl)))
    Trick = TrickUrlSession()
    Trick.setUrl(geturl)
    get_shibie_request = Trick.get(url=geturl, headers=headers)
    #print(get_shibie_request.json())
    get_shibie_result = get_shibie_request.json()
    shibie_guiji_res = get_shibie_result.get("results") if get_shibie_result.get("results") else group_id
    log.info("获取识别结果返回数据:--{}-- ".format(str(get_shibie_result)))
    log.info("获取识别结果输出数据:--{}-- ".format(str(shibie_guiji_res)))
    return shibie_guiji_res


#excel数据写入
def excel_write_open(sheet,title):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=sheet, title=title)
    return ws,wb

#excel写入新sheet
def excel_write_load(filename,sheet,title):
    wb = openpyxl.load_workbook(filename + '.xlsx')
    ws = wb.create_sheet(index=sheet, title=title)
    return ws,wb


#excel数据保存退出
def excel_write_save(filename,wb):
    return wb.save(filename + '.xlsx')

def timeutil(info_time):
    return time.strftime("%Y--%m--%d %H:%M:%S",time.localtime(int(info_time))) if len(str(info_time)) == 10 else int(time.mktime(time.strptime(info_time, "%Y-%m-%d %H:%M:%S")))

#渠道风控数据写入
async def qudao_cleaning(store_id,qudaolist,headers,ws):
    this_store_id = store_id
    ws.append(["客户姓名","首次到访时间","刷证时间","id_number","刷证图片","匹配的personid","person首次到访","相似度","底图","到访轨迹","到访时间","抓拍位置"])
    if isinstance(qudaolist,list):
        #返回的有渠道列表数据
        qudao_cursor = 0
        qudao_list_length = len(qudaolist)
        while qudao_cursor <= 30:
            if qudao_cursor == qudao_list_length:break
            if qudaolist[qudao_cursor].get('first_capture_time') != 0 and qudaolist[qudao_cursor].get('verify_time') != 0:
                #到访时间和刷证时间都有,进入获取渠道详情
                id_number = qudaolist[qudao_cursor].get('id_number')
                qudao_name = qudaolist[qudao_cursor].get("name")
                first_daofan = timeutil(qudaolist[qudao_cursor].get('first_capture_time'))
                verify_time = timeutil(qudaolist[qudao_cursor].get('verify_time'))
                id_capture_image,qudao_detail_res = get_qudao_detail_page(id_number=id_number,store_id=this_store_id,headers=headers)
                #print("id_capture_image",id_capture_image,"qudao_detail_res",qudao_detail_res)
                if isinstance(qudao_detail_res, list):
                    #print("详情内部搜图结果",qudao_detail_res)
                    for qudao_inner in qudao_detail_res:
                        person_id_search = qudao_inner.get('person_id')
                        arrived_at_search = qudao_inner.get('arrived_at')
                        confidence_search = qudao_inner.get('confidence')
                        arrived_image_url_search = qudao_inner.get('arrived_image_url')
                        trace = qudao_inner.get('trace')
                        #print("获取到的trace",trace)
                        #返回的tance是一个列表
                        inner = 0
                        trace_count = len(trace)
                        while inner <= 5:
                            if inner == trace_count:break
                            device_id = trace[inner].get('device_id')
                            camera_id = trace[inner].get('camera_id')
                            image_url = trace[inner].get('image_url')
                            #print("内存获取的图片地址:",image_url)
                            ts = trace[inner].get('ts')
                            camera_name = trace[inner].get('camera_name')
                            if inner == 0:
                                first_insert = [qudao_name,first_daofan,verify_time,id_number,id_capture_image,person_id_search,
                                                arrived_at_search,confidence_search,arrived_image_url_search,image_url,ts,camera_name]
                                ws.append(first_insert)
                            else:
                                already_insert = ['','','','','',person_id_search,
                                                arrived_at_search,confidence_search,arrived_image_url_search,image_url,ts,camera_name]
                                ws.append(already_insert)
                            inner += 1
                else:
                    temp_no_number = [qudao_name,first_daofan,verify_time,id_number,id_capture_image,"渠道风控详情搜索结果是空的"]
                    ws.append(temp_no_number)
            else:
                qudao_name = qudaolist[qudao_cursor].get('name')
                qudao_id_number = qudaolist[qudao_cursor].get('id_number')
                temp_no_number = [qudao_name,'','',qudao_id_number,"没有详情数据"]
                ws.append(temp_no_number)
            qudao_cursor += 1
    else:
        temp_result = [store_id,'','',"无渠道风控数据"]
        ws.append(temp_result)
    return "渠道风控数收集完毕!"


#自主报备数据写入
async def baobei_cleaning(store_id,baobeilist,headers,ws):
    this_store_id = store_id
    ws.append(["客户名称","身份证号","渠道","-id","首次到访","匹配的personid","person首次到访","相似度","底图","到访轨迹","到访时间","抓拍位置"])
    if isinstance(baobeilist,list):
        #返回的有渠道列表数据
        baobei_cursor = 0
        baobei_list_length = len(baobeilist)
        while baobei_cursor <= 30:
            if baobei_cursor == baobei_list_length:break
            if baobeilist[baobei_cursor].get('first_arrive') != 0:
                #有到访时间进入获取轨迹
                _id = baobeilist[baobei_cursor].get('_id')
                baofan_name = baobeilist[baobei_cursor].get('name')
                daofan_id_number = baobeilist[baobei_cursor].get('id_number')
                daofan_ditch = baobeilist[baobei_cursor].get('ditch')
                first_arrive = baobeilist[baobei_cursor].get('first_arrive')
                daofan_first_arrive = timeutil(first_arrive)
                print("匹配的首次到访时间",daofan_first_arrive)
                id_detail_image,baobei_detail_res = get_baobei_detail_page(store_id=this_store_id,_id=_id,headers=headers)
                print("自主报备返回的详情数据",baobei_detail_res)
                #print("id_capture_image",id_capture_image,"qudao_detail_res",qudao_detail_res)
                if isinstance(baobei_detail_res, list):
                    print("_______进入判断逻辑",isinstance(baobei_detail_res, list))
                    #print("自主报备详情内部搜图结果",baobei_detail_res)
                    for baobei_inner in baobei_detail_res:
                        person_id_search = baobei_inner.get('person_id')
                        arrived_at_search = baobei_inner.get('arrived_at')
                        confidence_search = baobei_inner.get('confidence')
                        arrived_image_url_search = baobei_inner.get('arrived_image_url')
                        trace = baobei_inner.get('trace')
                        #print("获取到的trace",trace)
                        #返回的tance是一个列表
                        inner = 0
                        trace_count = len(trace)
                        while inner <= 5:
                            if inner == trace_count:break
                            #device_id = trace[inner].get('device_id')
                            #camera_id = trace[inner].get('camera_id')
                            image_url = trace[inner].get('image_url')
                            print("内存获取的图片地址:",image_url)
                            ts = trace[inner].get('ts')
                            camera_name = trace[inner].get('camera_name')
                            if inner == 0:
                                first_insert = [baofan_name,daofan_id_number,daofan_ditch,_id,daofan_first_arrive,person_id_search,
                                                arrived_at_search,confidence_search,arrived_image_url_search,image_url,ts,camera_name]
                                ws.append(first_insert)
                            else:
                                already_insert = ['','','','','',person_id_search,
                                                arrived_at_search,confidence_search,arrived_image_url_search,image_url,ts,camera_name]
                                ws.append(already_insert)
                            inner += 1
                else:
                    temp_no_number = [baofan_name, daofan_id_number, daofan_ditch, _id, daofan_first_arrive, "详情匹配结果是空的"]
                    ws.append(temp_no_number)
            else:
                _id = baobeilist[baobei_cursor].get('_id')
                daofan_name = baobeilist[baobei_cursor].get('name')
                daofan_id_number = baobeilist[baobei_cursor].get('id_number')
                daofan_ditch = baobeilist[baobei_cursor].get('ditch')
                first_arrive = baobeilist[baobei_cursor].get('first_arrive')
                daofan_first_arrive = timeutil(first_arrive)
                #print("没有详情的数据首次到访记录",daofan_first_arrive)
                temp_no_number = [daofan_name,daofan_id_number,daofan_ditch,_id,daofan_first_arrive,"没有详情数据"]
                ws.append(temp_no_number)
            baobei_cursor += 1
    else:
        temp_result = [store_id,'','',"无自主报备数据"]
        ws.append(temp_result)
    return "自主报备数据收集完毕!"


#识别结果数据写入
async def shibie_cleaning(store_id,store_device,headers,ws):
    this_store_id = store_id
    write_title_temp = ['案场group','face_id','person_id','camera_name','request_id','底图地址','性别','年龄','轨迹地址','抓拍时间']
    ws.append(write_title_temp)
    if isinstance(store_device, str):
        write_temp = [store_id,store_device,"没有绑定设备!"]
        ws.append(write_temp)
        return store_id,"没有绑定设备"
    if isinstance(store_device,list):
        #绑定的有设备
        device_cursor = 0
        store_device_length = len(store_device)
        while device_cursor <= store_device_length:
            if device_cursor == store_device_length:break
            group_id = store_device[device_cursor].get('group_id')
            #store_id = store_device[device_cursor].get('store_id')
            device_id = store_device[device_cursor].get('device_id')
            created = store_device[device_cursor].get('created')
            device_cursor += 1
        #默认查询识别结果
        default_group_id = group_id
        default_device_id = device_id
        default_start_time = timeutil("2020-04-01 00:00:00")
        default_end_time = timeutil("2020-04-05 20:00:00")
        get_shibie_result = get_shibie_detail(group_id=default_group_id,start_time=default_start_time,end_time=default_end_time,device_id=default_device_id,headers=headers)
        print("默认时间获取的识别结果是：",get_shibie_result)
        if isinstance(get_shibie_result, list):
            #while循环读取识别结果数据
            shibie_face_cursor = 0
            shibie_face_count = len(get_shibie_result)
            while shibie_face_cursor < shibie_face_count:
                if shibie_face_cursor == shibie_face_count:break
                write_face_id = get_shibie_result[shibie_face_cursor].get("face_id")
                write_camera_name = get_shibie_result[shibie_face_cursor].get("camera_name")
                write_person_id = get_shibie_result[shibie_face_cursor].get("person_id")
                write_gender = get_shibie_result[shibie_face_cursor].get("gender")
                write_age = get_shibie_result[shibie_face_cursor].get("age")
                write_base_image_url = get_shibie_result[shibie_face_cursor].get("base_image_url")
                write_image_url = get_shibie_result[shibie_face_cursor].get("image_url")
                write_request_id = get_shibie_result[shibie_face_cursor].get("request_id")
                get_request_time = get_shibie_result[shibie_face_cursor].get("request_time")
                write_request_time = timeutil(get_request_time)
                write_group_id = get_shibie_result[shibie_face_cursor].get("group_id")
                if shibie_face_cursor == 0:
                    temp_write = [write_group_id,write_face_id,write_person_id,write_camera_name,write_request_id,write_base_image_url,write_gender,write_age,write_image_url,write_request_time]
                else:
                    temp_write = ["",write_face_id,write_person_id,write_camera_name,write_request_id,write_base_image_url,write_gender,write_age,write_image_url,write_request_time]
                ws.append(temp_write)
                shibie_face_cursor += 1
        else:
            #时间段没有识别结果，从开始时间开始查询识别结果
            shibie_start_time,shibie_end_time = get_shibie_time(store_id=this_store_id,start_time=created,headers=headers)
            print("计算过的有识别结果时间是",shibie_start_time,shibie_end_time)
            print("获取到的设备是",device_id)
            second_start_time = shibie_start_time
            second_end_time = shibie_end_time
            second_shibie_result = get_shibie_detail(group_id=default_group_id, start_time=second_start_time,
                                                  end_time=second_end_time, device_id=default_device_id,
                                                  headers=headers)
            print("时间后获取的识别结果是",second_shibie_result)
            second_face_cursor = 0
            second_face_count = len(second_shibie_result)
            while second_face_cursor < second_face_count:
                second_write_face_id = second_shibie_result[second_face_cursor].get("face_id")
                second_write_camera_name = second_shibie_result[second_face_cursor].get("camera_name")
                second_write_person_id = second_shibie_result[second_face_cursor].get("person_id")
                second_write_gender = second_shibie_result[second_face_cursor].get("gender")
                second_write_age = second_shibie_result[second_face_cursor].get("age")
                second_write_base_image_url = second_shibie_result[second_face_cursor].get("base_image_url")
                second_write_image_url = second_shibie_result[second_face_cursor].get("image_url")
                second_write_request_id = second_shibie_result[second_face_cursor].get("request_id")
                second_get_request_time = second_shibie_result[second_face_cursor].get("request_time")
                second_write_request_time = timeutil(second_get_request_time)
                second_write_group_id = second_shibie_result[second_face_cursor].get("group_id")
                if second_face_cursor == 0:
                    second_temp_write = [second_write_group_id, second_write_face_id, second_write_person_id, second_write_camera_name, second_write_request_id,
                                  second_write_base_image_url, second_write_gender, second_write_age, second_write_image_url, second_write_request_time]
                else:
                    second_temp_write = ["", second_write_face_id, second_write_person_id, second_write_camera_name, second_write_request_id,
                                  second_write_base_image_url, second_write_gender, second_write_age, second_write_image_url, second_write_request_time]
                ws.append(second_temp_write)
                second_face_cursor += 1
        return "识别结果数据收集完毕!"


#以图搜图-写入保存图片
async def image_save_shibie(store_id,store_name,store_device,headers):
    aims_path = os.path.join(os.path.dirname(__file__), store_name)
    if not os.path.exists(aims_path):
        os.mkdir(aims_path)
    #2、判断案场是否绑定设备
    if isinstance(store_device, str):
        print(store_id, store_device, "没有绑定设备!")
        return store_id, store_device
    if isinstance(store_device, list):
        # 绑定的有设备
        res = requests.session()
        device_cursor = 0
        store_device_length = len(store_device)
        while device_cursor < store_device_length:
            if device_cursor == store_device_length:break
            group_id = store_device[device_cursor].get('group_id')
            device_id = store_device[device_cursor].get('device_id')
            created = store_device[device_cursor].get('created')
            device_cursor += 1
        #查询识别结果
        default_group_id = group_id
        default_device_id = device_id
        default_start_time = timeutil("2020-04-01 00:00:00")
        default_end_time = timeutil("2020-04-05 20:00:00")
        get_shibie_result = get_shibie_detail(group_id=default_group_id, start_time=default_start_time,
                                              end_time=default_end_time, device_id=default_device_id, headers=headers)
        #print("默认时间获取的识别结果是：", get_shibie_result)
        if isinstance(get_shibie_result, list):
            # 识别结果中取图片
            shibie_face_cursor = 0
            shibie_face_count = len(get_shibie_result)
            image_person_id = []
            while shibie_face_cursor < shibie_face_count:
                if shibie_face_cursor > 100:break
                image_person_id_info = get_shibie_result[shibie_face_cursor].get("person_id")
                image_image_url = get_shibie_result[shibie_face_cursor].get("image_url")
                if image_person_id_info in image_person_id:
                    print("personid已经存在跳过图片保存")
                    shibie_face_cursor += 1
                    continue
                else:
                    image_person_id.append(image_person_id_info)
                    image_result = res.get(url=image_image_url)
                    image_temp = image_person_id_info + '.jpg'
                    image_path = os.path.join(aims_path,image_temp)
                    with open(image_path,'wb') as fp:
                        fp.write(image_result.content)
                shibie_face_cursor += 1
        else:
            # 时间段没有识别结果，从开始时间开始查询识别结果
            shibie_start_time, shibie_end_time = get_shibie_time(store_id=store_id, start_time=created,
                                                                 headers=headers)
            print("计算过的有识别结果时间是", shibie_start_time, shibie_end_time)
            print("获取到的设备是", device_id)
            second_start_time = shibie_start_time
            second_end_time = shibie_end_time
            second_shibie_result = get_shibie_detail(group_id=default_group_id, start_time=second_start_time,
                                                     end_time=second_end_time, device_id=default_device_id,
                                                     headers=headers)
            print("时间后获取的识别结果是", second_shibie_result)
            second_face_cursor = 0
            second_face_count = len(second_shibie_result)
            second_image_person_id = []
            while second_face_cursor < second_face_count:
                if second_face_cursor > 100:break
                second_write_person_id = second_shibie_result[second_face_cursor].get("person_id")
                second_write_image_url = second_shibie_result[second_face_cursor].get("image_url")
                if second_write_person_id in second_image_person_id:
                    print("person已经存在过！")
                    second_face_cursor += 1
                    continue
                else:
                    second_image_person_id.append(second_write_person_id)
                    image_result = res.get(second_write_image_url)
                    print(image_result.content)
                    second_image_result = res.get(url=second_write_image_url)
                    image_temp = second_write_person_id + '.jpg'
                    image_path = os.path.join(aims_path, image_temp)
                    with open(image_path, 'wb') as fp:
                        fp.write(second_image_result.content)
                second_face_cursor += 1
        return store_id,"图片已经保存完成！"


def image_utils(store_id):
    #读文件夹下图片
    aims_path = os.path.join(os.path.dirname(__file__), store_id)
    if os.path.exists(aims_path):
        temp = os.listdir(aims_path)
        image_cursor = 0
        temp_length = len(temp)
        res_image_path_list = []
        person_image_id = []
        while image_cursor < temp_length:
            image_temp = temp[image_cursor]
            if image_temp.endswith('.jpg'):
                aims_image_path = os.path.join(aims_path, image_temp)
                with open(aims_image_path, "rb") as fp:
                    base64_data = base64.b64encode(fp.read())
                    base64_str_data = base64_data.decode()
                res_image_path_list.append(base64_str_data)
                person_image_id.append(image_temp.replace('.jpg',''))
            image_cursor += 1
        #print("目录存在返回结果",res_image_path_list)
        return res_image_path_list,person_image_id
    else:
        temp = os.listdir(aims_path)
        image_cursor = 0
        temp_length = len(temp)
        res_image_path_list = []
        person_image_id = []
        while image_cursor < temp_length:
            image_temp = temp[image_cursor]
            if image_temp.endswith('.jpg'):
                aims_image_path = os.path.join(aims_path,image_temp)
                with open(aims_image_path, "rb") as fp:
                    base64_data = base64.b64encode(fp.read())
                    base64_str_data = base64_data.decode()
                res_image_path_list.append(base64_str_data)
                person_image_id.append(image_temp.replace('.jpg', ''))
            image_cursor += 1
        print("目录不存在返回结果",res_image_path_list)
        return res_image_path_list,person_image_id


#以图搜图——搜图结果写入
async def search_image_cleaning(group_id,face_image_list,person_image_id,headers,ws):
    #搜图写入图片
    search_cursor = 0
    search_count = len(face_image_list)
    ws.append(['group_id', '图片personid', '匹配personid', '首次到访时间', '相似度', '图片地址', '设备名称', '轨迹图片', '时间', '抓拍位置'])
    while search_cursor < search_count:
        if search_cursor == search_count:break
        face_image_data = face_image_list[search_cursor]
        face_image_search_personid = person_image_id[search_cursor]
        search_image_list = await get_search_image_page(group_id,face_image_data,headers)
        search_cursor += 1
        if isinstance(search_image_list, list):
            # 返回的有渠道列表数据
            search_image_cursor = 0
            search_image_list_length = len(search_image_list)
            first_search_person_id = [group_id,face_image_search_personid,'','','','','','','','']
            ws.append(first_search_person_id)
            while search_image_cursor < search_image_list_length:
                if search_image_cursor == search_image_list_length:break
                search_person_id = search_image_list[search_image_cursor].get('person_id')
                search_arrived_at = search_image_list[search_image_cursor].get('arrived_at')
                search_camera_name = search_image_list[search_image_cursor].get('camera_name')
                search_confidence = search_image_list[search_image_cursor].get('confidence')
                search_arrived_image_url = search_image_list[search_image_cursor].get('arrived_image_url')
                trace = search_image_list[search_image_cursor].get('trace')
                # 返回的tance是一个列表
                inner = 0
                trace_count = len(trace)
                while inner < 30:
                    if inner == trace_count:break
                    inner_image_url = trace[inner].get('image_url')
                    ts = trace[inner].get('ts')
                    camera_name = trace[inner].get('camera_name')
                    if inner == 0:
                        first_insert = ['','',search_person_id,search_arrived_at,search_confidence,
                                        search_arrived_image_url,search_camera_name,inner_image_url,ts, camera_name]
                        print("首次写入",first_insert)
                        ws.append(first_insert)
                    else:
                        already_insert = ['','','','','',search_arrived_image_url,search_camera_name,inner_image_url,ts, camera_name]
                        print("后续数据写入",already_insert)
                        ws.append(already_insert)
                    inner += 1
                search_image_cursor += 1
        else:
            temp_no_number = [group_id,face_image_search_personid,'', '没有搜索到结果！']
            print("没有搜索到轨迹",temp_no_number)
            ws.append(temp_no_number)
    return "以图搜图数据收集完毕!"


#yaml配置文件读取
def config_yaml():
    curpath = os.path.dirname(os.path.realpath(__file__))
    yamlpath = os.path.join(curpath, 'config.yaml')
    f = open(yamlpath, 'r', encoding='utf-8')
    cfg = f.read()
    # 解决安全警告配置形成字典
    d = yaml.load(cfg, Loader=yaml.FullLoader)
    yaml.warnings({'YAMLLoadWarning': False})
    return d


#主函数
async def main():
    #参数配置
    config_dict = config_yaml()
    print(config_dict)
    username = config_dict['username']
    password = config_dict['password']
    store_id = config_dict['store_id']
    print(store_id,"类型是",type(store_id))
    #获取请求token
    url = "https://icloud.sensetime.com/senserealty/login"
    #login_res = await token_main(url=url,username=username,password=password)   #1.登录主函数获取用户名密码
    #token = login_res.get("token")
    #companyid = login_res.get("companyid")
    #组装保持会话headers
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36",
        "token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiSUQxNzI3IiwiY29tcGFueV9pZCI6IklEMTE3NCIsInVzZXJfdHlwZSI6ImFkbWluIiwibm9kZV9pZCI6IjAwMCIsImFrIjoibDEtOGRhMDViM2UtbHM1NmQ4OGYxZmZmIiwidXNlcl9hZ2VudCI6MSwiZXhwIjoxNTg2MjI4OTQ1fQ.JswP8CH8q1yyLWEzPfCuFEKbXNzSg1isBBq-JP17T9s",#eval(token),
        "Referer": "https://icloud.sensetime.com/senserealty/store/operate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Pragma": "no-cache",
        "Host": "icloud.sensetime.com",
        "Accept": "application/json, text/plain, */*"
        }
    #得到案场名称和案场store_id字典数据
    storeres = await store_list_find('ID1174',headers)               #eval(companyid) 传递的会带引号需要去掉
    print(storeres,":",type(storeres))
    store_id_name = get_store_name(**storeres)             #根据案场id字典取值方式获取中文名
    print(store_id_name,":",type(store_id_name))
    #获取所有案场id和中文名关系后开始循环写入数据
    store_id_write_len = len(store_id)
    store_write_info_cursor = 0
    while store_write_info_cursor < store_id_write_len:
        #定义文件夹
        write_store_id = store_id[store_write_info_cursor]
        write_store_name = store_id_name[write_store_id]
        write_aims_path = os.path.join(os.path.join(os.path.dirname(__file__), write_store_name),write_store_name)
        aims_path = os.path.join(os.path.dirname(__file__), write_store_name)
        print("目标路径",write_aims_path)
        if not os.path.exists(aims_path):
            os.mkdir(aims_path)
        # 1.渠道风控数据写入
        store_id_qudao,qudaonum = await get_qudao_page(write_store_id,headers)   #1.先调用渠道方法返回渠道数据
        #返回案场id表示没有渠道风控数据
        ws,wb = excel_write_open(0,"渠道风控数据")
        qudao_result = await qudao_cleaning(store_id_qudao,qudaonum,headers,ws)      #2.整和写入渠道详情数据
        print(qudao_result)
        excel_write_save(write_aims_path,wb)
        #2.自主报备数据写入(自主报备可以不刷证，返回身份证号可以为空)
        store_id_baobei, baobeinum = await get_baobei_page(write_store_id,headers)
        ws,wb = excel_write_load(write_aims_path,1,"自主报备数据")
        baobei_result = await baobei_cleaning(store_id_baobei,baobeinum,headers,ws)
        print(baobei_result)
        excel_write_save(write_aims_path,wb)
        #3.识别结果数据写入
        store_id_shibie,group_device_map = await get_shibie_store_device(write_store_id,headers)
        ws, wb = excel_write_load(write_aims_path, 2, "识别结果数据")
        shibie_result = await shibie_cleaning(store_id_shibie,group_device_map,headers,ws)
        print(shibie_result)
        excel_write_save(write_aims_path, wb)
        #4.以图搜图
        store_id_shibie,group_device_map = await get_shibie_store_device(write_store_id,headers)
        #aims_path = os.path.join(os.path.dirname(__file__),write_store_name)
        ws, wb = excel_write_load(write_aims_path, 3, "以图搜图数据")
        temp_list = os.listdir(aims_path)
        flag = 0
        for i in temp_list:
            if i.endswith('.jpg'):
                flag = 1
                break
        print(flag)
        if not flag:
            image_write_store_id,write_result = await image_save_shibie(store_id_shibie,write_store_name,group_device_map,headers)
            print(image_write_store_id,write_result)
            search_image_list,person_image_id = image_utils(write_store_name)
            search_image_result = await search_image_cleaning(store_id_shibie,search_image_list,person_image_id,headers,ws)
            print("if分之搜图返回结果:",search_image_result)
        else:
            search_image_list,person_image_id = image_utils(write_store_name)
            search_image_result = await search_image_cleaning(store_id_shibie, search_image_list,person_image_id, headers,ws)
            print("else分之搜图返回结果:", search_image_result)
        excel_write_save(write_aims_path,wb)
        store_write_info_cursor += 1


if __name__=="__main__":
    #asyncio.get_event_loop().run_until_complete(main())
    #asyncio.get_event_loop().run_until_complete(store_find())
    #asyncio.get_event_loop().run_until_complete(qudao_page('cdb978497a'))
    #asyncio.get_event_loop().run_until_complete(qudao_detail())
    #asyncio.get_event_loop().run_until_complete(baobei_page())
    #asyncio.get_event_loop().run_until_complete(baobei_detail())
    #asyncio.get_event_loop().run_until_complete(search_page())
    #asyncio.get_event_loop().run_until_complete(daofang_page_count())
    #asyncio.get_event_loop().run_until_complete(daofang_page_summary())
    #asyncio.get_event_loop().run_until_complete(shibie_detail())
    #asyncio.get_event_loop().run_until_complete(shibie_detail())
    #image_utils('a14c29a517')
    asyncio.get_event_loop().run_until_complete(main())