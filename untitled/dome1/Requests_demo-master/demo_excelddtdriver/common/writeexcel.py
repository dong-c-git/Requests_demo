#coding:utf-8
from openpyxl import load_workbook
import openpyxl

def copy_excel(excelpath1,excelpath2):
    """
    复制excel,把excelpath1的数据复制到excelpath2中
    :param excelpath1:
    :param excelpath2:
    :return:
    """
    print(excelpath2)
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)

    #读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row         #最大行数
    max_column = sheet1.max_column   #最大列数

    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):  #chr(97)="a"
            n = chr(n)
            i = "%s%d"%(n,m)
            cell1 = sheet1[i].value
            sheet2[i].value = cell1
    wb2.save(excelpath2)
    wb1.close()
    wb2.close()

class Write_excel(object):
    """
    修改excel数据
    """
    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  #激活sheet

    def write(self,row_n,col_n,value):
        """
        写入数据，如(2,3,"hello") 第二行第三列写入hello
        :param row_n:
        :param col_n:
        :param value:
        :return:
        """
        self.ws.cell(row_n,col_n).value = value
        self.wb.save(self.filename)

if __name__=="__main__":
    copy_excel("debug_api.xlsx",str("testreport.xlsx"))
    # wt = Write_excel("test111.xlsx")
    # wt.write(4,5,"HELLEOP")
    # wt.write(4,6,"HELLEOP")
    # wb2 = openpyxl.Workbook()
    # # 指定当前显示（活动）的sheet对象
    # ws = wb2.active
    # # 给A1单元格赋值
    # ws['A1'] = 42
    # # 一行添加多列数据
    # ws.append([1, 2, 3])
    # # 保存excel
    # wb2.save("sample.xlsx")
